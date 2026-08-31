# tools.py — Data fetching layer for the financial agent

from yahooquery import search
import yfinance as yf
import feedparser
import requests
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
# ─────────────────────────────────────────
# ← NEW: DOCUMENT ANALYSIS TOOLS
# ─────────────────────────────────────────
import PyPDF2              # ← NEW import
import docx               # ← NEW import
import openpyxl          # ← NEW import
from PIL import Image    # ← NEW import
import pytesseract       # ← NEW import
# Add this line right after your imports so Python knows where the OCR engine lives
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

import difflib
import os


load_dotenv()

def resolve_stock_ticker(stock_name: str):

    stock_name = stock_name.strip().upper()

    if stock_name.endswith(".NS"):
        return stock_name

    try:
        results = search(stock_name)

        for quote in results.get("quotes", []):

            symbol = quote.get("symbol", "")
            exchange = quote.get("exchange", "")

            if symbol.endswith(".NS"):
                return symbol

            if exchange == "NSI":
                return symbol + ".NS"

        return None

    except Exception:
        return None
# ─────────────────────────────────────────
# TOOL 1 — Get current stock data
# ─────────────────────────────────────────
def get_stock_data(stock_input: str) -> dict:
    """
    Fetches current price and key metrics.

    Accepts:
    - Company name (Adani Power)
    - NSE ticker (ADANIPOWER)
    - Yahoo ticker (ADANIPOWER.NS)

    Automatically resolves to the correct Yahoo Finance ticker.
    """

    try:
        # Resolve to Yahoo ticker
        ticker = resolve_stock_ticker(stock_input)

        if ticker is None:
            return {
                "error": f"Could not find stock '{stock_input}'."
            }

        stock = yf.Ticker(ticker)
        info = stock.info

        if not info or info.get("longName") is None:
            return {
                "error": f"No market data found for '{stock_input}'."
            }

        return {
            "ticker": ticker,
            "company_name": info.get("longName", "N/A"),
            "current_price": info.get("currentPrice", "N/A"),
            "previous_close": info.get("previousClose", "N/A"),
            "day_high": info.get("dayHigh", "N/A"),
            "day_low": info.get("dayLow", "N/A"),
            "52w_high": info.get("fiftyTwoWeekHigh", "N/A"),
            "52w_low": info.get("fiftyTwoWeekLow", "N/A"),
            "volume": info.get("volume", "N/A"),
            "market_cap": info.get("marketCap", "N/A"),
            "pe_ratio": info.get("trailingPE", "N/A"),
            "sector": info.get("sector", "N/A"),
            "industry": info.get("industry", "N/A"),
        }

    except Exception as e:
        return {"error": str(e)}


# ─────────────────────────────────────────
# TOOL 2 — Get historical price data
# ─────────────────────────────────────────
def get_historical_data(ticker: str, period: str = "3mo") -> str:
    """
    Fetches historical OHLCV data.
    period options: 1mo, 3mo, 6mo, 1y, 2y
    """
    try:
        stock = yf.Ticker(ticker)
        df = stock.history(period=period)
        df = df[["Open", "High", "Low", "Close", "Volume"]]
        df.index = df.index.strftime("%Y-%m-%d")
        return df.tail(10).to_string()
    except Exception as e:
        return f"Error: {str(e)}"


# ─────────────────────────────────────────
# TOOL 3 — Get latest news for a stock
# Source: NewsAPI (free tier: 100 req/day)
# ─────────────────────────────────────────
# ─────────────────────────────────────────
# Helper function for news insights
# ─────────────────────────────────────────
def generate_news_insight(headline: str, company: str) -> str:
    """
    Creates a brief insight from the news headline including impact sentiment.
    """
    headline_lower = headline.lower()
    company_lower = company.lower()
    
    positive_keywords = ["surge", "jump", "rise", "gain", "profit", "approval", "growth", "expansion", "strong", "beat", "rally", "soars"]
    negative_keywords = ["fall", "drop", "decline", "loss", "halt", "crash", "weak", "miss", "investigation", "warning", "plunge", "tumble"]
    
    positive_count = sum(1 for keyword in positive_keywords if keyword in headline_lower)
    negative_count = sum(1 for keyword in negative_keywords if keyword in headline_lower)
    
    # Determine sentiment impact
    if positive_count > negative_count:
        sentiment = "✅ Positive Impact"
    elif negative_count > positive_count:
        sentiment = "⚠️ Negative Impact"
    else:
        sentiment = "➡️ Neutral Impact"
    
    # Categorize news type
    if any(word in headline_lower for word in ["dividend", "bonus", "buyback"]):
        category = "Shareholder returns announcement"
    elif any(word in headline_lower for word in ["profit", "earnings", "revenue"]):
        category = "Financial performance update"
    elif any(word in headline_lower for word in ["approval", "clearance", "nod"]):
        category = "Regulatory approval received"
    elif any(word in headline_lower for word in ["expansion", "launch", "new"]):
        category = "Growth/expansion initiative"
    elif any(word in headline_lower for word in ["merger", "acquisition", "stake"]):
        category = "M&A or stake activity"
    else:
        category = headline[:60] + "..." if len(headline) > 60 else headline
    
    # Return combined insight
    insight = f"{sentiment} — {category}"
    return insight
def get_stock_news(query: str) -> list[dict]:
    """
    Fetch news for a stock using Google News RSS feed.
    Works for any time period (no 30-day limit).
    """
    try:
        import feedparser
        
        # Google News RSS feed for stock query
        url = f"https://news.google.com/rss/search?q={query}+stock&hl=en-US&gl=US&ceid=US:en"
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        feed = feedparser.parse(url, request_headers=headers)
        
        if not feed.entries:
            return [{"error": f"No news found for {query}"}]
        
        results = []
        for entry in feed.entries[:5]:  # Top 5 articles
            results.append({
                "title": entry.get('title', 'N/A'),
                "source": entry.get('source', {}).get('title', 'Google News'),
                "published": entry.get('published', 'Unknown'),
                "description": entry.get('summary', 'N/A')[:200],  # Short summary
                "url": entry.get('link', '')
            })
        
        return results
        
    except Exception as e:
        return [{"error": f"Failed to fetch news: {str(e)}"}]

# ─────────────────────────────────────────
# TOOL 4 — Get sector relationships
# ─────────────────────────────────────────
def get_sector_impact(sector: str) -> dict:
    """
    Returns which sectors/factors directly and indirectly
    impact a given sector. Built-in knowledge base.
    """
    sector_map = {
        "Technology": {
            "direct_impacts": ["USD/INR exchange rate", "US tech spending", "IT hiring trends"],
            "indirect_impacts": ["Global inflation", "Fed interest rates", "Startup ecosystem"],
            "key_indicators": ["NASDAQ movement", "US GDP growth", "Digital transformation budgets"]
        },
        "Energy": {
            "direct_impacts": ["Crude oil prices", "Government energy policy", "Refining margins"],
            "indirect_impacts": ["Geopolitical tensions", "INR/USD rate", "EV adoption rate"],
            "key_indicators": ["Brent crude price", "Natural gas prices", "Renewable energy targets"]
        },
        "Banking": {
            "direct_impacts": ["RBI interest rates", "Credit growth", "NPA levels"],
            "indirect_impacts": ["Inflation", "GDP growth", "Government spending"],
            "key_indicators": ["Repo rate", "Credit-deposit ratio", "Bond yields"]
        },
        "Automobile": {
            "direct_impacts": ["Steel prices", "Fuel prices", "Consumer sentiment"],
            "indirect_impacts": ["Interest rates on loans", "Rural income", "EV policy"],
            "key_indicators": ["Monthly sales data", "Two-wheeler sales", "EV registrations"]
        },
        "Pharmaceuticals": {
            "direct_impacts": ["US FDA approvals", "API raw material costs", "USD/INR rate"],
            "indirect_impacts": ["Healthcare policy", "Patent cliffs", "R&D spending"],
            "key_indicators": ["USFDA inspection outcomes", "Generic drug pricing", "Export data"]
        }
    }

    default = {
        "direct_impacts": ["Market sentiment", "Sector-specific regulations", "Input costs"],
        "indirect_impacts": ["Global macro trends", "INR/USD movement", "Government policy"],
        "key_indicators": ["Nifty sector index", "FII/DII flows", "Quarterly earnings"]
    }

    return sector_map.get(sector, default)


# ─────────────────────────────────────────
# TOOL 5 — Get Mutual Fund NAV data
# ─────────────────────────────────────────
# ─────────────────────────────────────────
# TOOL 5 — Get Mutual Fund NAV data
# Source: MFAPI.in (free, no API key needed)
# ─────────────────────────────────────────
import requests
import pandas as pd
from datetime import datetime

def get_mutual_fund_nav(fund_name: str) -> list[dict]:
    """
    Fetches the current, 3-year ago, and 5-year ago NAV data from MFAPI.in.
    Accepts a partial name (e.g., "SBI Bluechip", "HDFC Midcap").
    """
    try:
        # STEP 1: Search the master list to find matching funds and their scheme codes
        master_url = "https://api.mfapi.in/mf"
        master_response = requests.get(master_url, timeout=30)
        
        if master_response.status_code != 200:
            return [{"error": "Failed to fetch master scheme list from MFAPI."}]
            
        all_schemes = master_response.json()
        
        # Filter schemes matching the partial fund name
        matched_schemes = [
            scheme for scheme in all_schemes 
            if fund_name.lower() in scheme["schemeName"].lower()
        ]
        
        if not matched_schemes:
            return [{"error": f"No funds found matching '{fund_name}'"}]
            
        # Limit to top 3 matches to avoid hitting rate limits in the next step
        results = []
        for scheme in matched_schemes[:3]:
            scheme_code = scheme["schemeCode"]
            scheme_name = scheme["schemeName"]
            
            # STEP 2: Fetch the full timeline history for this specific scheme code
            history_url = f"https://api.mfapi.in/mf/{scheme_code}"
            history_response = requests.get(history_url, timeout=30)
            
            if history_response.status_code != 200:
                continue
                
            history_data = history_response.json().get("data", [])
            if not history_data:
                continue
                
            # Convert history to a Pandas DataFrame for reliable date parsing and looking back
            df = pd.DataFrame(history_data)
            df['date'] = pd.to_datetime(df['date'], format='%d-%m-%Y')
            df['nav'] = pd.to_numeric(df['nav'])
            df.set_index('date', inplace=True)
            df.sort_index(ascending=True, inplace=True) # Sort oldest to newest

            # Define time horizons
            today = pd.Timestamp.now()
            target_3y = today - pd.DateOffset(years=3)
            target_5y = today - pd.DateOffset(years=5)
            
            # Extract latest NAV (last item in chronological order)
            current_date = df.index[-1]
            current_nav = df.iloc[-1]['nav']
            
            # Extract historical NAVs using .asof() to handle weekends and market holidays intelligently
            # If the exact date was a Sunday, it safely grabs the Friday right before it.
            try:
                nav_3y_record = df.asof(target_3y)
                nav_3y = float(nav_3y_record['nav']) if not pd.isna(nav_3y_record['nav']) else None
                date_3y = df.index[df.index.get_indexer([target_3y], method='pad')].strftime('%d-%b-%Y')
            except Exception:
                nav_3y, date_3y = "Data Unavailable", "N/A"
                
            try:
                nav_5y_record = df.asof(target_5y)
                nav_5y = float(nav_5y_record['nav']) if not pd.isna(nav_5y_record['nav']) else None
                date_5y = df.index[df.index.get_indexer([target_5y], method='pad')].strftime('%d-%b-%Y')
            except Exception:
                nav_5y, date_5y = "Data Unavailable", "N/A"

            results.append({
                "scheme_name": scheme_name,
                "scheme_code": scheme_code,
                "current_nav": current_nav,
                "current_date": current_date.strftime('%d-%b-%Y'),
                "nav_3_years_ago": nav_3y,
                "date_3_years_ago": date_3y,
                "nav_5_years_ago": nav_5y,
                "date_5_years_ago": date_5y
            })
            
        return results if results else [{"error": "Could not extract history for matched funds."}]

    except Exception as e:
        return [{"error": str(e)}]

# ─────────────────────────────────────────
# TOOL 6 — Get Economic Indicators
# ─────────────────────────────────────────
def get_economic_indicators() -> dict:
    """
    Fetches key Indian economic indicators from World Bank API.
    Free, no API key needed, official source.
    """
    try:
        indicators = {
            "GDP Growth (%)": "NY.GDP.MKTP.KD.ZG",
            "Inflation (%)": "FP.CPI.TOTL.ZG",
            "Unemployment (%)": "SL.UEM.TOTL.ZS",
            "Current Account Balance (% GDP)": "BN.CAB.XOKA.GD.ZS"
        }

        results = {}
        base_url = "https://api.worldbank.org/v2/country/IN/indicator"

        for name, code in indicators.items():
            url = f"{base_url}/{code}?format=json&mrv=1"
            response = requests.get(url, timeout=30)
            data = response.json()

            if len(data) > 1 and data[1]:
                latest = data[1][0]
                results[name] = {
                    "value": round(latest["value"], 2) if latest["value"] else "N/A",
                    "year": latest["date"]
                }
            else:
                results[name] = {"value": "N/A", "year": "N/A"}

        return results

    except Exception as e:
        return {"error": str(e)}


# ─────────────────────────────────────────
# TOOL 7 — Get Competitor Comparison
# ─────────────────────────────────────────
def get_competitor_data(ticker: str) -> dict:
    """
    Fetches and compares key metrics for a stock and its competitors.
    ticker: e.g. "TCS.NS", "RELIANCE.NS", "HDFCBANK.NS"
    """
    try:
        competitor_map = {
            "TCS.NS":       ["TCS.NS", "INFY.NS", "WIPRO.NS", "HCLTECH.NS"],
            "INFY.NS":      ["INFY.NS", "TCS.NS", "WIPRO.NS", "HCLTECH.NS"],
            "WIPRO.NS":     ["WIPRO.NS", "TCS.NS", "INFY.NS", "HCLTECH.NS"],
            "HCLTECH.NS":   ["HCLTECH.NS", "TCS.NS", "INFY.NS", "WIPRO.NS"],
            "RELIANCE.NS":  ["RELIANCE.NS", "ONGC.NS", "IOC.NS", "BPCL.NS"],
            "HDFCBANK.NS":  ["HDFCBANK.NS", "ICICIBANK.NS", "SBIN.NS", "AXISBANK.NS"],
            "ICICIBANK.NS": ["ICICIBANK.NS", "HDFCBANK.NS", "SBIN.NS", "AXISBANK.NS"],
            "SBIN.NS":      ["SBIN.NS", "HDFCBANK.NS", "ICICIBANK.NS", "AXISBANK.NS"],
            "MARUTI.NS":    ["MARUTI.NS", "TATAMOTORS.NS", "M&M.NS", "BAJAJ-AUTO.NS"],
            "SUNPHARMA.NS": ["SUNPHARMA.NS", "DRREDDY.NS", "CIPLA.NS", "DIVISLAB.NS"],
        }

        peers = competitor_map.get(ticker, [ticker])
        comparison = []

        for peer in peers:
            stock = yf.Ticker(peer)
            info = stock.info
            comparison.append({
                "ticker": peer,
                "company": info.get("longName", "N/A"),
                "price": info.get("currentPrice", "N/A"),
                "pe_ratio": info.get("trailingPE", "N/A"),
                "market_cap": info.get("marketCap", "N/A"),
                "52w_high": info.get("fiftyTwoWeekHigh", "N/A"),
                "52w_low": info.get("fiftyTwoWeekLow", "N/A"),
                "revenue_growth": info.get("revenueGrowth", "N/A"),
                "dividend_yield": round(info.get("dividendYield", 0) * 100, 2) if info.get("dividendYield") else "N/A",
            })

        return {"competitors": comparison}

    except Exception as e:
        return {"error": str(e)}


# ─────────────────────────────────────────
# TOOL 8 — Get Price Alerts
# ─────────────────────────────────────────
def get_price_alerts(ticker: str) -> dict:
    """
    Flags if stock is near 52 week high or low.
    ticker: NSE symbol like RELIANCE.NS
    """
    try:
        stock = yf.Ticker(ticker)
        info = stock.info

        current = info.get("currentPrice", 0)
        high_52w = info.get("fiftyTwoWeekHigh", 0)
        low_52w = info.get("fiftyTwoWeekLow", 0)

        if not current or not high_52w or not low_52w:
            return {"alert": "Data unavailable"}

        from_high = round(((current - high_52w) / high_52w) * 100, 2)
        from_low = round(((current - low_52w) / low_52w) * 100, 2)

        alerts = []

        if from_low <= 5:
            alerts.append(f"🔴 CRITICAL: Trading within 5% of 52 week low (₹{low_52w}) — strong support zone or breakdown risk")
        elif from_low <= 15:
            alerts.append(f"🟡 WARNING: Trading within 15% of 52 week low (₹{low_52w}) — watch for further downside")

        if from_high >= -5:
            alerts.append(f"🟢 BREAKOUT ZONE: Trading within 5% of 52 week high (₹{high_52w}) — bullish momentum")
        elif from_high >= -15:
            alerts.append(f"🟡 NOTE: Trading within 15% of 52 week high (₹{high_52w}) — recovering well")

        if not alerts:
            alerts.append(f"📊 NEUTRAL: Trading in middle range — {from_low}% above 52w low, {abs(from_high)}% below 52w high")

        return {
            "current_price": current,
            "52w_high": high_52w,
            "52w_low": low_52w,
            "from_52w_high": f"{from_high}%",
            "from_52w_low": f"+{from_low}%",
            "alerts": alerts
        }

    except Exception as e:
        return {"error": str(e)}
# ─────────────────────────────────────────
# TOOL 9 — Get Historical Growth %
# ─────────────────────────────────────────
def get_historical_growth(ticker: str) -> dict:
    """
    Fetches 3 year and 5 year price growth % for a stock.
    ticker: NSE symbol like TCS.NS, RELIANCE.NS
    """
    try:
        stock = yf.Ticker(ticker)
        
        # Get 5 years of historical data
        df = stock.history(period="5y")
        
        if df.empty:
            return {"error": "No historical data available"}
        
        current_price = df["Close"].iloc[-1]
        
        # 3 year growth
        three_year_ago_index = -756  # approx 252 trading days per year
        if len(df) >= 756:
            price_3y_ago = df["Close"].iloc[three_year_ago_index]
            growth_3y = round(((current_price - price_3y_ago) / price_3y_ago) * 100, 2)
        else:
            growth_3y = "N/A"

        # 5 year growth
        if len(df) >= 1260:
            price_5y_ago = df["Close"].iloc[0]
            growth_5y = round(((current_price - price_5y_ago) / price_5y_ago) * 100, 2)
        else:
            growth_5y = "N/A"

        return {
            "ticker": ticker,
            "current_price": round(current_price, 2),
            "3yr_growth_%": growth_3y,
            "5yr_growth_%": growth_5y
        }

    except Exception as e:
        return {"error": str(e)}

# ─────────────────────────────────────────
# TOOL 10 — Get Mutual Fund Holdings
# Top 10 Indian funds — update figures from
# Valueresearch.in or AMC websites monthly
# ─────────────────────────────────────────
def get_fund_holdings(fund_name: str) -> dict:
    """
    Returns top 6 holdings with % allocation for top 10 Indian funds.
    fund_name: e.g. "SBI Bluechip", "Parag Parikh Flexi Cap"
    """
    holdings_map = {

       # ==========================================
    # ── 1. ORIGINAL 10 UPDATED FUNDS ──
    # ==========================================
    "parag parikh flexi cap": {
        "fund_name": "Parag Parikh Flexi Cap Fund",
        "category": "Flexi Cap",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 8.12},
            {"stock": "Power Grid Corporation", "percentage": 6.85},
            {"stock": "Bajaj Holdings & Investment", "percentage": 6.40},
            {"stock": "ITC", "percentage": 5.90},
            {"stock": "Coal India", "percentage": 5.25},
            {"stock": "Alphabet Inc (Google)", "percentage": 4.80},
            {"stock": "Maruti Suzuki India", "percentage": 4.35},
        ],
    },
    "hdfc flexi cap": {
        "fund_name": "HDFC Flexi Cap Fund",
        "category": "Flexi Cap",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 8.69},
            {"stock": "Axis Bank", "percentage": 6.83},
            {"stock": "HDFC Bank", "percentage": 6.81},
            {"stock": "State Bank of India", "percentage": 4.74},
            {"stock": "Larsen & Toubro", "percentage": 3.35},
            {"stock": "Bharti Airtel", "percentage": 3.10},
            {"stock": "Reliance Industries", "percentage": 2.19},
        ],
    },
    "sbi bluechip": {
        "fund_name": "SBI Bluechip Fund",
        "category": "Large Cap",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 8.65},
            {"stock": "ICICI Bank", "percentage": 7.80},
            {"stock": "Reliance Industries", "percentage": 6.40},
            {"stock": "Larsen & Toubro", "percentage": 5.45},
            {"stock": "ITC", "percentage": 4.10},
            {"stock": "Infosys", "percentage": 3.85},
            {"stock": "Axis Bank", "percentage": 3.30},
        ],
    },
    "mirae asset large cap": {
        "fund_name": "Mirae Asset Large Cap Fund",
        "category": "Large Cap",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 9.25},
            {"stock": "ICICI Bank", "percentage": 8.10},
            {"stock": "Reliance Industries", "percentage": 6.70},
            {"stock": "Infosys", "percentage": 5.05},
            {"stock": "TCS", "percentage": 3.95},
            {"stock": "Larsen & Toubro", "percentage": 3.60},
            {"stock": "Axis Bank", "percentage": 3.35},
        ],
    },
    "hdfc midcap opportunities": {
        "fund_name": "HDFC Midcap Opportunities Fund",
        "category": "Mid Cap",
        "holdings": [
            {"stock": "Cholamandalam Investment", "percentage": 4.25},
            {"stock": "Max Healthcare Institute", "percentage": 3.90},
            {"stock": "Persistent Systems", "percentage": 3.75},
            {"stock": "Supreme Industries", "percentage": 3.35},
            {"stock": "Indian Hotels Company", "percentage": 3.10},
            {"stock": "Crompton Greaves Consumer", "percentage": 2.85},
            {"stock": "Tube Investments of India", "percentage": 2.75},
        ],
    },
    "axis midcap": {
        "fund_name": "Axis Midcap Fund",
        "category": "Mid Cap",
        "holdings": [
            {"stock": "Cholamandalam Investment", "percentage": 4.55},
            {"stock": "Persistent Systems", "percentage": 3.85},
            {"stock": "Trent", "percentage": 3.60},
            {"stock": "Page Industries", "percentage": 3.40},
            {"stock": "Astral", "percentage": 3.05},
            {"stock": "Tube Investments of India", "percentage": 2.90},
            {"stock": "Divi's Laboratories", "percentage": 2.75},
        ],
    },
    "sbi small cap": {
        "fund_name": "SBI Small Cap Fund",
        "category": "Small Cap",
        "holdings": [
            {"stock": "Finolex Cables", "percentage": 4.10},
            {"stock": "Blue Star", "percentage": 3.85},
            {"stock": "Hawkins Cookers", "percentage": 3.55},
            {"stock": "Kalpataru Projects", "percentage": 3.45},
            {"stock": "Elgi Equipments", "percentage": 3.30},
            {"stock": "Garware Technical Fibres", "percentage": 3.10},
            {"stock": "Suprajit Engineering", "percentage": 2.90},
        ],
    },
    "nippon india small cap": {
        "fund_name": "Nippon India Small Cap Fund",
        "category": "Small Cap",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 1.89},
            {"stock": "Bharat Heavy Electricals", "percentage": 1.67},
            {"stock": "TD Power Systems", "percentage": 1.54},
            {"stock": "Apar Industries", "percentage": 1.49},
            {"stock": "Karur Vysya Bank", "percentage": 1.29},
            {"stock": "State Bank of India", "percentage": 1.18},
            {"stock": "Tube Investments of India", "percentage": 1.01},
        ],
    },
    "mirae asset tax saver": {
        "fund_name": "Mirae Asset ELSS Tax Saver Fund",
        "category": "ELSS",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 8.85},
            {"stock": "ICICI Bank", "percentage": 7.35},
            {"stock": "Reliance Industries", "percentage": 6.25},
            {"stock": "Infosys", "percentage": 4.75},
            {"stock": "TCS", "percentage": 3.85},
            {"stock": "Larsen & Toubro", "percentage": 3.40},
            {"stock": "Axis Bank", "percentage": 3.20},
        ],
    },
    "uti nifty 50 index": {
        "fund_name": "UTI Nifty 50 Index Fund",
        "category": "Index Fund",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 11.35},
            {"stock": "Reliance Industries", "percentage": 9.45},
            {"stock": "ICICI Bank", "percentage": 7.90},
            {"stock": "Infosys", "percentage": 5.15},
            {"stock": "TCS", "percentage": 3.75},
            {"stock": "Larsen & Toubro", "percentage": 3.60},
            {"stock": "ITC", "percentage": 3.45},
        ],
    },
    # ==========================================
    # ── 2. ADDITIONAL 50 KEY FUNDS ──
    # ==========================================
    # --- FLEXI CAP ---
    "jm flexi cap": {
        "fund_name": "JM Flexi Cap Fund",
        "category": "Flexi Cap",
        "holdings": [
            {"stock": "Larsen & Toubro", "percentage": 5.40},
            {"stock": "NTPC", "percentage": 4.80},
            {"stock": "ICICI Bank", "percentage": 4.50},
            {"stock": "REC Ltd", "percentage": 4.10},
            {"stock": "State Bank of India", "percentage": 3.90},
        ],
    },
    "kotak flexi cap": {
        "fund_name": "Kotak Flexi Cap Fund",
        "category": "Flexi Cap",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 7.20},
            {"stock": "HDFC Bank", "percentage": 6.80},
            {"stock": "Larsen & Toubro", "percentage": 5.10},
            {"stock": "Infosys", "percentage": 4.30},
            {"stock": "Reliance Industries", "percentage": 3.80},
        ],
    },
    "motilal oswal flexi cap": {
        "fund_name": "Motilal Oswal Flexi Cap Fund",
        "category": "Flexi Cap",
        "holdings": [
            {"stock": "Jio Financial Services", "percentage": 8.50},
            {"stock": "Zomato", "percentage": 6.40},
            {"stock": "Trent", "percentage": 5.90},
            {"stock": "Larsen & Toubro", "percentage": 5.10},
            {"stock": "ICICI Bank", "percentage": 4.70},
        ],
    },
    # --- LARGE CAP ---
    "icici prudential bluechip": {
        "fund_name": "ICICI Prudential Bluechip Fund",
        "category": "Large Cap",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 9.10},
            {"stock": "Reliance Industries", "percentage": 8.20},
            {"stock": "HDFC Bank", "percentage": 7.80},
            {"stock": "Larsen & Toubro", "percentage": 5.20},
            {"stock": "Infosys", "percentage": 4.50},
        ],
    },
    "nippon india large cap": {
        "fund_name": "Nippon India Large Cap Fund",
        "category": "Large Cap",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 9.30},
            {"stock": "ICICI Bank", "percentage": 8.40},
            {"stock": "ITC", "percentage": 5.60},
            {"stock": "Reliance Industries", "percentage": 5.20},
            {"stock": "State Bank of India", "percentage": 4.10},
        ],
    },
    "canara robeco bluechip": {
        "fund_name": "Canara Robeco Bluechip Equity Fund",
        "category": "Large Cap",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 9.40},
            {"stock": "ICICI Bank", "percentage": 8.50},
            {"stock": "Infosys", "percentage": 5.20},
            {"stock": "Reliance Industries", "percentage": 4.90},
            {"stock": "Larsen & Toubro", "percentage": 3.80},
        ],
    },
    # --- LARGE & MID CAP ---
    "kotak equity opportunities": {
        "fund_name": "Kotak Equity Opportunities Fund",
        "category": "Large & Mid Cap",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 5.90},
            {"stock": "HDFC Bank", "percentage": 5.20},
            {"stock": "SRF", "percentage": 3.80},
            {"stock": "Trent", "percentage": 3.40},
            {"stock": "Larsen & Toubro", "percentage": 3.10},
        ],
    },
    "mirae asset large & midcap": {
        "fund_name": "Mirae Asset Large & Midcap Fund",
        "category": "Large & Mid Cap",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 5.80},
            {"stock": "ICICI Bank", "percentage": 5.10},
            {"stock": "State Bank of India", "percentage": 3.50},
            {"stock": "Axis Bank", "percentage": 3.20},
            {"stock": "NTPC", "percentage": 2.80},
        ],
    },
    "sbi large & midcap": {
        "fund_name": "SBI Large & Midcap Fund",
        "category": "Large & Mid Cap",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 4.90},
            {"stock": "Reliance Industries", "percentage": 4.20},
            {"stock": "Page Industries", "percentage": 3.10},
            {"stock": "Infosys", "percentage": 2.90},
            {"stock": "Trent", "percentage": 2.70},
        ],
    },
    # --- MID CAP ---
    "motilal oswal midcap": {
        "fund_name": "Motilal Oswal Midcap Fund",
        "category": "Mid Cap",
        "holdings": [
            {"stock": "Jio Financial Services", "percentage": 7.20},
            {"stock": "Zomato", "percentage": 6.80},
            {"stock": "Trent", "percentage": 5.40},
            {"stock": "Coforge", "percentage": 4.90},
            {"stock": "Persistent Systems", "percentage": 4.30},
        ],
    },
    "nippon india growth": {
        "fund_name": "Nippon India Growth Fund",
        "category": "Mid Cap",
        "holdings": [
            {"stock": "Power Finance Corp", "percentage": 3.90},
            {"stock": "Cholamandalam Investment", "percentage": 3.50},
            {"stock": "Supreme Industries", "percentage": 3.20},
            {"stock": "Varun Beverages", "percentage": 2.90},
            {"stock": "BWD Industries", "percentage": 2.60},
        ],
    },
    "kotak emerging equity": {
        "fund_name": "Kotak Emerging Equity Fund",
        "category": "Mid Cap",
        "holdings": [
            {"stock": "Supreme Industries", "percentage": 4.10},
            {"stock": "Persistent Systems", "percentage": 3.80},
            {"stock": "Cummins India", "percentage": 3.40},
            {"stock": "Schaeffler India", "percentage": 3.00},
            {"stock": "Solar Industries", "percentage": 2.80},
        ],
    },
    # --- SMALL CAP ---
    "quant small cap": {
        "fund_name": "Quant Small Cap Fund",
        "category": "Small Cap",
        "holdings": [
            {"stock": "Reliance Industries", "percentage": 6.10},
            {"stock": "Jio Financial Services", "percentage": 4.80},
            {"stock": "Aegis Logistics", "percentage": 3.90},
            {"stock": "HFCL", "percentage": 3.20},
            {"stock": "Bikaji Foods", "percentage": 2.90},
        ],
    },
    "axis small cap": {
        "fund_name": "Axis Small Cap Fund",
        "category": "Small Cap",
        "holdings": [
            {"stock": "Narayana Hrudayalaya", "percentage": 3.90},
            {"stock": "Brigade Enterprises", "percentage": 3.50},
            {"stock": "CCL Products", "percentage": 3.20},
            {"stock": "Krishna Institute of Medical", "percentage": 2.90},
            {"stock": "Chalet Hotels", "percentage": 2.80},
        ],
    },
    "tata small cap": {
        "fund_name": "Tata Small Cap Fund",
        "category": "Small Cap",
        "holdings": [
            {"stock": "IDFC First Bank", "percentage": 3.40},
            {"stock": "BASF India", "percentage": 3.10},
            {"stock": "Quess Corp", "percentage": 2.80},
            {"stock": "Allcargo Logistics", "percentage": 2.60},
            {"stock": "DCB Bank", "percentage": 2.50},
        ],
    },
    # --- MULTI CAP ---
    "nippon india multi cap": {
        "fund_name": "Nippon India Multi Cap Fund",
        "category": "Multi Cap",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 4.80},
            {"stock": "ICICI Bank", "percentage": 4.20},
            {"stock": "Larsen & Toubro", "percentage": 3.10},
            {"stock": "Linde India", "percentage": 2.60},
            {"stock": "Reliance Industries", "percentage": 2.40},
        ],
    },
    "mahindra manulife multi cap": {
        "fund_name": "Mahindra Manulife Multi Cap Fund",
        "category": "Multi Cap",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 4.50},
            {"stock": "ICICI Bank", "percentage": 3.90},
            {"stock": "Infosys", "percentage": 2.80},
            {"stock": "Larsen & Toubro", "percentage": 2.60},
            {"stock": "Tata Motors", "percentage": 2.30},
        ],
    },
    "icici prudential multicap": {
        "fund_name": "ICICI Prudential Multicap Fund",
        "category": "Multi Cap",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 5.10},
            {"stock": "Infosys", "percentage": 3.80},
            {"stock": "HDFC Bank", "percentage": 3.60},
            {"stock": "Reliance Industries", "percentage": 3.20},
            {"stock": "TCS", "percentage": 2.70},
        ],
    },
    # --- ELSS ---
    "quant elss tax saver": {
        "fund_name": "Quant ELSS Tax Saver Fund",
        "category": "ELSS",
        "holdings": [
            {"stock": "Reliance Industries", "percentage": 7.40},
            {"stock": "Jio Financial Services", "percentage": 5.20},
            {"stock": "Adani Power", "percentage": 4.30},
            {"stock": "Hindalco Industries", "percentage": 3.80},
            {"stock": "Steel Authority of India", "percentage": 3.10},
        ],
    },
    "dsp elss tax saver": {
        "fund_name": "DSP ELSS Tax Saver Fund",
        "category": "ELSS",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 6.80},
            {"stock": "HDFC Bank", "percentage": 6.20},
            {"stock": "Infosys", "percentage": 4.50},
            {"stock": "Axis Bank", "percentage": 3.90},
            {"stock": "Larsen & Toubro", "percentage": 3.40},
        ],
    },
    "sbi long term equity": {
        "fund_name": "SBI Long Term Equity Fund",
        "category": "ELSS",
        "holdings": [
            {"stock": "GE T&D India", "percentage": 4.80},
            {"stock": "ICICI Bank", "percentage": 4.50},
            {"stock": "Reliance Industries", "percentage": 3.90},
            {"stock": "Mahindra & Mahindra", "percentage": 3.20},
            {"stock": "Bharti Airtel", "percentage": 2.90},
        ],
    },
    # --- VALUE / CONTRA ---
    "sbi contra": {
        "fund_name": "SBI Contra Fund",
        "category": "Value/Contra",
        "holdings": [
            {"stock": "GAIL India", "percentage": 3.80},
            {"stock": "HDFC Bank", "percentage": 3.40},
            {"stock": "Cognizant Technology Solutions", "percentage": 3.10},
            {"stock": "State Bank of India", "percentage": 2.90},
            {"stock": "Oil & Natural Gas Corp", "percentage": 2.70},
        ],
    },
    "icici prudential value discovery": {
        "fund_name": "ICICI Prudential Value Discovery Fund",
        "category": "Value/Contra",
        "holdings": [
            {"stock": "NTPC", "percentage": 8.20},
            {"stock": "ICICI Bank", "percentage": 6.90},
            {"stock": "Bharti Airtel", "percentage": 5.40},
            {"stock": "ONGC", "percentage": 4.80},
            {"stock": "Sun Pharma", "percentage": 3.90},
        ],
    },
    "invesco india contra": {
        "fund_name": "Invesco India Contra Fund",
        "category": "Value/Contra",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 7.10},
            {"stock": "HDFC Bank", "percentage": 6.40},
            {"stock": "Infosys", "percentage": 4.30},
            {"stock": "Reliance Industries", "percentage": 3.90},
            {"stock": "NTPC", "percentage": 3.20},
        ],
    },
    # --- FOCUSED ---
    "360 one focused equity": {
        "fund_name": "360 ONE Focused Equity Fund",
        "category": "Focused",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 9.80},
            {"stock": "HDFC Bank", "percentage": 8.90},
            {"stock": "Infosys", "percentage": 7.20},
            {"stock": "Larsen & Toubro", "percentage": 6.50},
            {"stock": "Bharti Airtel", "percentage": 5.90},
        ],
    },
    "sbi focused equity": {
        "fund_name": "SBI Focused Equity Fund",
        "category": "Focused",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 8.90},
            {"stock": "ICICI Bank", "percentage": 7.40},
            {"stock": "Divi's Laboratories", "percentage": 6.20},
            {"stock": "TCS", "percentage": 5.80},
            {"stock": "Bajaj Finance", "percentage": 5.10},
        ],
    },
    "hdfc focused 30": {
        "fund_name": "HDFC Focused 30 Fund",
        "category": "Focused",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 9.10},
            {"stock": "HDFC Bank", "percentage": 8.40},
            {"stock": "Axis Bank", "percentage": 6.50},
            {"stock": "NTPC", "percentage": 5.20},
            {"stock": "Larsen & Toubro", "percentage": 4.60},
        ],
    },
    # --- SECTORAL / THEMATIC ---
    "icici prudential infrastructure": {
        "fund_name": "ICICI Prudential Infrastructure Fund",
        "category": "Sectoral/Thematic",
        "holdings": [
            {"stock": "Larsen & Toubro", "percentage": 11.20},
            {"stock": "NTPC", "percentage": 8.40},
            {"stock": "Bharti Airtel", "percentage": 6.90},
            {"stock": "Gujarat Gas", "percentage": 4.50},
            {"stock": "Oil & Natural Gas Corp", "percentage": 4.10},
        ],
    },
    "tata digital india": {
        "fund_name": "Tata Digital India Fund",
        "category": "Sectoral/Thematic",
        "holdings": [
            {"stock": "Infosys", "percentage": 18.50},
            {"stock": "TCS", "percentage": 14.20},
            {"stock": "HCL Technologies", "percentage": 9.80},
            {"stock": "Tech Mahindra", "percentage": 7.40},
            {"stock": "Persistent Systems", "percentage": 5.60},
        ],
    },
    "sbi healthcare opportunities": {
        "fund_name": "SBI Healthcare Opportunities Fund",
        "category": "Sectoral/Thematic",
        "holdings": [
            {"stock": "Sun Pharma", "percentage": 15.20},
            {"stock": "Cipla", "percentage": 10.40},
            {"stock": "Dr. Reddy's Laboratories", "percentage": 8.90},
            {"stock": "Divi's Laboratories", "percentage": 7.60},
            {"stock": "Apollo Hospitals", "percentage": 6.10},
        ],
    },
    "nippon india power & infra": {
        "fund_name": "Nippon India Power & Infra Fund",
        "category": "Sectoral/Thematic",
        "holdings": [
            {"stock": "Reliance Industries", "percentage": 8.90},
            {"stock": "Larsen & Toubro", "percentage": 7.50},
            {"stock": "NTPC", "percentage": 6.80},
            {"stock": "Power Grid Corp", "percentage": 5.90},
            {"stock": "Tata Power", "percentage": 4.80},
        ],
    },
    # --- BALANCED ADVANTAGE ---
    "icici prudential balanced advantage": {
        "fund_name": "ICICI Prudential Balanced Advantage Fund",
        "category": "Balanced Advantage",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 6.50},
            {"stock": "HDFC Bank", "percentage": 5.40},
            {"stock": "TV Derivatives / Hedged Position", "percentage": 18.50},
            {"stock": "Government Securities (Debt)", "percentage": 22.10},
            {"stock": "Reliance Industries", "percentage": 3.20},
        ],
    },
    "hdfc balanced advantage": {
        "fund_name": "HDFC Balanced Advantage Fund",
        "category": "Balanced Advantage",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 5.90},
            {"stock": "ICICI Bank", "percentage": 5.10},
            {"stock": "Coal India", "percentage": 4.10},
            {"stock": "Government Securities (Debt)", "percentage": 28.40},
            {"stock": "NTPC", "percentage": 3.20},
        ],
    },
    "edelweiss balanced advantage": {
        "fund_name": "Edelweiss Balanced Advantage Fund",
        "category": "Balanced Advantage",
        "holdings": [
            {"stock": "ICICI Bank", "percentage": 5.20},
            {"stock": "HDFC Bank", "percentage": 4.60},
            {"stock": "Government Securities (Debt)", "percentage": 31.50},
            {"stock": "TCS", "percentage": 2.90},
            {"stock": "Infosys", "percentage": 2.70},
        ],
    },
    # --- MULTI ASSET ALLOCATION ---
    "icici prudential multi-asset": {
        "fund_name": "ICICI Prudential Multi-Asset Fund",
        "category": "Multi Asset",
        "holdings": [
            {"stock": "ICICI Bank (Equity)", "percentage": 5.80},
            {"stock": "ICICI Prudential Gold ETF", "percentage": 11.40},
            {"stock": "ICICI Prudential Silver ETF", "percentage": 4.20},
            {"stock": "Government Debt Securities", "percentage": 24.50},
            {"stock": "NTPC (Equity)", "percentage": 3.90},
        ],
    },
    "nippon india multi asset": {
        "fund_name": "Nippon India Multi Asset Allocation Fund",
        "category": "Multi Asset",
        "holdings": [
            {"stock": "Nippon India ETF Gold BeES", "percentage": 12.80},
            {"stock": "HDFC Bank (Equity)", "percentage": 4.80},
            {"stock": "Government Securities (Debt)", "percentage": 21.30},
            {"stock": "Reliance Industries (Equity)", "percentage": 3.90},
            {"stock": "Commodity Derivatives / Silver", "percentage": 5.10},
        ],
    },
    "quant multi asset": {
        "fund_name": "Quant Multi Asset Fund",
        "category": "Multi Asset",
        "holdings": [
            {"stock": "Reliance Industries (Equity)", "percentage": 6.90},
            {"stock": "Gold ETFs", "percentage": 14.20},
            {"stock": "Treasury Bills (Debt)", "percentage": 18.50},
            {"stock": "Jio Financial Services", "percentage": 4.10},
            {"stock": "Silver ETFs", "percentage": 3.80},
        ],
    },
    # --- ARBITRAGE ---
    "kotak equity arbitrage": {
        "fund_name": "Kotak Equity Arbitrage Fund",
        "category": "Arbitrage",
        "holdings": [
            {"stock": "Cash-Futures Hedged Arbitrage", "percentage": 68.50},
            {"stock": "TREPS / Clearing Corp", "percentage": 18.20},
            {"stock": "Commercial Papers", "percentage": 8.40},
            {"stock": "Corporate Fixed Deposits", "percentage": 4.90},
        ],
    },
    "invesco india arbitrage": {
        "fund_name": "Invesco India Arbitrage Fund",
        "category": "Arbitrage",
        "holdings": [
            {"stock": "Hedging Arbitrage Spread", "percentage": 67.10},
            {"stock": "TREPS / Repo Market", "percentage": 20.40},
            {"stock": "Bank Certificates of Deposit", "percentage": 7.80},
            {"stock": "Treasury Bills", "percentage": 4.70},
        ],
    },
    # --- DEBT / LIQUID / SHORT DURATION ---
    "hdfc liquid": {
        "fund_name": "HDFC Liquid Fund",
        "category": "Debt/Liquid/Short Duration",
        "holdings": [
            {"stock": "91 Days Treasury Bills", "percentage": 28.40},
            {"stock": "182 Days Treasury Bills", "percentage": 18.20},
            {"stock": "NABARD Commercial Paper", "percentage": 7.50},
            {"stock": "TREPS / Reverse Repo", "percentage": 12.10},
            {"stock": "HDFC Bank CD", "percentage": 6.80},
        ],
    },
    "icici prudential savings": {
        "fund_name": "ICICI Prudential Savings Fund",
        "category": "Debt/Liquid/Short Duration",
        "holdings": [
            {"stock": "NABARD Corporate Bonds", "percentage": 8.90},
            {"stock": "REC Ltd Bonds", "percentage": 7.40},
            {"stock": "PFC Corporate Bonds", "percentage": 6.80},
            {"stock": "182 Days Treasury Bills", "percentage": 14.50},
            {"stock": "TREPS", "percentage": 9.20},
        ],
    },
    "sbi short term debt": {
        "fund_name": "SBI Short Term Debt Fund",
        "category": "Debt/Liquid/Short Duration",
        "holdings": [
            {"stock": "Government Securities (G-Sec)", "percentage": 34.20},
            {"stock": "NABARD AAA Bonds", "percentage": 9.80},
            {"stock": "PFC AAA Bonds", "percentage": 8.10},
            {"stock": "TREPS / Liquid Cash", "percentage": 7.50},
            {"stock": "Small Industries Dev Bank", "percentage": 6.40},
        ],
    },
    "aditya birla sun life corporate bond": {
        "fund_name": "Aditya Birla Sun Life Corporate Bond Fund",
        "category": "Debt/Liquid/Short Duration",
        "holdings": [
            {"stock": "REC Ltd AAA Bonds", "percentage": 9.50},
            {"stock": "NABARD AAA Bonds", "percentage": 8.80},
            {"stock": "PFC AAA Bonds", "percentage": 8.20},
            {"stock": "Government Securities", "percentage": 21.40},
            {"stock": "L&T Infra Bonds", "percentage": 5.90},
        ],
    },
    # --- INDEX FUNDS ---
    "navi nifty 50 index": {
        "fund_name": "Navi Nifty 50 Index Fund",
        "category": "Index Fund",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 11.35},
            {"stock": "Reliance Industries", "percentage": 9.45},
            {"stock": "ICICI Bank", "percentage": 7.90},
            {"stock": "Infosys", "percentage": 5.15},
            {"stock": "TCS", "percentage": 3.75},
        ],
    },
    "motilal oswal nifty midcap 150 index": {
        "fund_name": "Motilal Oswal Nifty Midcap 150 Index Fund",
        "category": "Index Fund",
        "holdings": [
            {"stock": "Max Healthcare Institute", "percentage": 2.25},
            {"stock": "Persistent Systems", "percentage": 2.10},
            {"stock": "Indian Hotels Company", "percentage": 1.95},
            {"stock": "Cummins India", "percentage": 1.80},
            {"stock": "Suzlon Energy", "percentage": 1.75},
        ],
    },
    "icici prudential nifty next 50 index": {
        "fund_name": "ICICI Prudential Nifty Next 50 Index Fund",
        "category": "Index Fund",
        "holdings": [
            {"stock": "Bharat Electronics", "percentage": 4.20},
            {"stock": "Trent", "percentage": 4.10},
            {"stock": "Tata Power", "percentage": 3.80},
            {"stock": "REC Ltd", "percentage": 3.50},
            {"stock": "Power Finance Corp", "percentage": 3.40},
        ],
    },
    "hdfc index sensex": {
        "fund_name": "HDFC Index Fund - Sensex Plan",
        "category": "Index Fund",
        "holdings": [
            {"stock": "HDFC Bank", "percentage": 13.10},
            {"stock": "Reliance Industries", "percentage": 11.20},
            {"stock": "ICICI Bank", "percentage": 9.20},
            {"stock": "Infosys", "percentage": 6.10},
            {"stock": "TCS", "percentage": 4.40},
        ],
    },
    # --- GOLD / INTERNATIONAL ---
    "sbi gold": {
        "fund_name": "SBI Gold Fund",
        "category": "Gold/International",
        "holdings": [
            {"stock": "SBI ETF Gold (Physical Gold)", "percentage": 98.20},
            {"stock": "TREPS / Cash Margins", "percentage": 1.80},
        ],
    },
    "franklin us opportunities": {
        "fund_name": "Franklin India Feeder - Franklin U.S. Opportunities",
        "category": "Gold/International",
        "holdings": [
            {"stock": "Microsoft Corp", "percentage": 8.90},
            {"stock": "NVIDIA Corp", "percentage": 8.20},
            {"stock": "Amazon.com Inc", "percentage": 6.80},
            {"stock": "Meta Platforms Inc", "percentage": 5.10},
            {"stock": "Apple Inc", "percentage": 4.60},
        ],
    },
    }

    # Search by partial name match
    query = fund_name.lower().strip()
    for key, data in holdings_map.items():
        if query in key or key in query:
            return data

    # If not found
    return {
        "error": f"Holdings not available for '{fund_name}'",
        "suggestion": "Please upload the fund factsheet PDF for holdings data"
    },

# ─────────────────────────────────────────
# Tool 13: Stock fundamentals
# ─────────────────────────────────────────
def get_stock_fundamentals(ticker: str):
    """
    Get key fundamental and financial-health metrics for an Indian stock.
    """
    try:
        import yfinance as yf

        if not ticker.endswith(".NS"):
            ticker = ticker + ".NS"

        stock = yf.Ticker(ticker)
        info = stock.info

        return {
            "ticker": ticker,
            "company": info.get("longName"),

            # Valuation
            "pe_ratio": info.get("trailingPE") or info.get("forwardPE") or 0,
            "pb_ratio": info.get("priceToBook"),
            "dividend_yield": info.get("dividendYield"),

            # Profitability
            "roe": info.get("returnOnEquity"),
            "roa": info.get("returnOnAssets"),
            "profit_margin": info.get("profitMargins"),
            "operating_margin": info.get("operatingMargins"),

            # Financial health
            "debt_to_equity": info.get("debtToEquity"),
            "current_ratio": info.get("currentRatio"),
            "quick_ratio": info.get("quickRatio"),

            # Debt / balance sheet
            "total_debt": info.get("totalDebt"),
            "total_cash": info.get("totalCash"),

            # Growth
            "revenue_growth": info.get("revenueGrowth"),
            "earnings_growth": info.get("earningsGrowth"),
        }

    except Exception as e:
        return {
            "error": str(e),
            "ticker": ticker
        },
# ─────────────────────────────────────────
# Tool 11: New document analysis
# ─────────────────────────────────────────
# ─────────────────────────────────────────
# DOCUMENT ANALYSIS TOOLS  --- updated
# ─────────────────────────────────────────

def extract_document_content(file_path: str) -> dict:
    """
    Extracts text content from various file types.
    Supports: PDF, DOCX, XLSX, PNG, JPG, JPEG, TXT
    """
    try:
        file_ext = file_path.lower().split('.')[-1]
        print(f"DEBUG extract_document_content: file_ext={file_ext}, path={file_path}")
        
        if file_ext == 'pdf':
            return extract_pdf(file_path)
        elif file_ext == 'docx':
            return extract_docx(file_path)
        elif file_ext == 'xlsx':
            return extract_xlsx(file_path)
        elif file_ext in ['png', 'jpg', 'jpeg']:
            return extract_image(file_path)
        elif file_ext == 'txt':
            return extract_txt(file_path)
        else:
            return {"error": f"Unsupported file type: .{file_ext}"}
    
    except Exception as e:
        print(f"DEBUG extract_document_content error: {str(e)}")
        return {"error": f"Document extraction failed: {str(e)}"}


def extract_pdf(file_path: str) -> dict:
    """Extract text from PDF"""
    try:
        import PyPDF2
        text = ""
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages[:50]:
                text += page.extract_text() or ""
        
        return {
            "content": text[:10000],
            "file_type": "pdf"
        }
    except Exception as e:
        return {"error": f"PDF extraction failed: {str(e)}"}


def extract_docx(file_path: str) -> dict:
    """Extract text from Word document"""
    try:
        import docx
        doc = docx.Document(file_path)
        text = "\n".join([p.text for p in doc.paragraphs])
        return {
            "content": text[:10000],
            "file_type": "docx"
        }
    except Exception as e:
        return {"error": f"DOCX extraction failed: {str(e)}"}


def extract_xlsx(file_path: str) -> dict:
    """Extract data from Excel"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        text = ""
        for sheet in wb.sheetnames[:3]:
            ws = wb[sheet]
            text += f"\n=== {sheet} ===\n"
            for row in list(ws.iter_rows(values_only=True))[:50]:
                text += str(row) + "\n"
        return {
            "content": text[:10000],
            "file_type": "xlsx"
        }
    except Exception as e:
        return {"error": f"XLSX extraction failed: {str(e)}"}


def extract_txt(file_path: str) -> dict:
    """Extract text from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            text = f.read()
        return {
            "content": text[:10000],
            "file_type": "txt"
        }
    except Exception as e:
        return {"error": f"TXT extraction failed: {str(e)}"}


def extract_image(file_path: str) -> dict:
    """Extract text from image using OCR"""
    try:
        from PIL import Image
        import pytesseract
        img = Image.open(file_path)
        text = pytesseract.image_to_string(img)
        return {
            "content": text[:10000],
            "file_type": "image"
        }
    except Exception as e:
        return {"error": f"Image extraction failed (try PNG/JPG): {str(e)}"}

# ─────────────────────────────────────────
# Tool 12: Recommendation tool by analysing investment worthiness
# ─────────────────────────────────────────
def analyze_investment_worthiness(ticker: str, fund_name: str = None) -> dict:
    """
    Comprehensive analysis for investment decision
    Returns: risk_profile, valuation, growth, sentiment
    """
    try:
        if ticker:
            # Get stock data
            stock_data = get_stock_data(ticker)
            growth_data = get_historical_growth(ticker)
            news_data = get_stock_news(ticker)
            
            return {
                "valuation": {
                    "pe_ratio": stock_data.get("pe_ratio", "N/A"),
                    "assessment": "Undervalued" if stock_data.get("pe_ratio", 999) < 15 else "Fairly valued" if stock_data.get("pe_ratio", 0) < 25 else "Overvalued"
                },
                "growth": {
                    "three_year_cagr": growth_data.get("three_year_cagr", "N/A"),
                    "trend": "Strong" if growth_data.get("three_year_cagr", 0) > 15 else "Moderate" if growth_data.get("three_year_cagr", 0) > 5 else "Weak"
                },
                "sentiment": {
                    "recent_news_count": len(news_data),
                    "sentiment": "Positive" if len(news_data) > 0 else "Neutral"
                },
                "risk_factors": {
                    "volatility": "High" if stock_data.get("pe_ratio", 0) > 25 else "Medium",
                    "sector_concentration": "Check if already invested in same sector"
                }
            }
    except Exception as e:
        return {"error": str(e)}

